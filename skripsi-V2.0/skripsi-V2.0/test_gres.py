from asyncore import read
from app import read_excel

from openpyxl import load_workbook

print(load_workbook('data\ModifiedS-Box.xlsx').sheetnames)