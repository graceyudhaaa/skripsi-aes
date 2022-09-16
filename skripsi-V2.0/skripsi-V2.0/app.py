import json, pyscrypt, time
import base64
from io import BytesIO
import numpy as np
from openpyxl import load_workbook
from PIL import Image
import pandas as pd
from flask import Flask, render_template, send_from_directory, request, abort, redirect, url_for, flash, g, send_file

from aes import AES
import aesoriginal

app = Flask(__name__)

app.secret_key = 'fCa7_l^6Uza@*uuh*->K)+%@eEZ+<z'
app.config['ENV'] = 'development'

#IMAGE PROCESSING

def img_to_rgb(img):
    img = Image.open(img)
    arr = np.array(img)

    return arr.flatten(), img.size, img.mode

def rgb_to_bytes(arr):
    return bytes(arr)

def bytes_to_image(b, mode, image_size):
    img = Image.frombytes(mode, image_size, b)

    return img

def image_to_base64(img):
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue())
    return img_str.decode()

def shannon_entropy(signal):
    lensig = signal.size
    symset = list(set(signal))
    numsym = len(symset)
    propab = [np.size(signal[signal==i])/(1.0*lensig) for i in symset]
    ent = np.sum([p*np.log2(1.0/p) for p in propab])

    return ent

def image_entropy(img, mode):
    signal = np.array(img).flatten()

    if mode == 'RGB':
        r = shannon_entropy(signal[0::3])
        g = shannon_entropy(signal[1::3])
        b = shannon_entropy(signal[2::3])

        avg = (r+g+b) / 3

        return avg
    else:
        return shannon_entropy(signal)

def UACI(img1, img2):
    pixel1 = img1.load()
    pixel2 = img2.load()

    width,height=img1.size
    value=0.0
    for y in range(0,height):
        for x in range(0,width):
            if type(pixel1[x,y]) == int:
                value=(abs(pixel1[x,y]-pixel2[x,y])/255)+value
            else:
                value=(abs(pixel1[x,y][0]-pixel2[x,y][0])/255)+value

    value=(value/(width*height))*100
    return value

def rateofchange(height,width,pixel1,pixel2,matrix,i):
    for y in range(0,height):
        for x in range(0,width):

            if type(pixel1[x,y]) == int:
                if pixel1[x,y] == pixel2[x,y]:
                    matrix[x,y]=0
                else:
                    matrix[x,y]=1
            else:
                if pixel1[x,y][i] == pixel2[x,y][i]:
                    matrix[x,y]=0
                else:
                    matrix[x,y]=1

    return matrix

def sumofpixel(height,width,pixel1,pixel2,ematrix,i):
    matrix=rateofchange(height,width,pixel1,pixel2,ematrix,i)
    psum=0
    for y in range(0,height):
        for x in range(0,width):
            psum=matrix[x,y]+psum
    return psum

def NPCR(img1,img2):
    width, height = img1.size
    pixel1 = img1.load()
    pixel2 = img2.load()
    ematrix = np.empty([width, height])
    per=(((sumofpixel(height,width,pixel1,pixel2,ematrix,0)/(height*width))*100)+((sumofpixel(height,width,pixel1,pixel2,ematrix,1)/(height*width))*100)+((sumofpixel(height,width,pixel1,pixel2,ematrix,2)/(height*width))*100))/3
    return per


#AES FUNCTION

SBOX = None

def read_excel(sbox_file, worksheet):

    wb = load_workbook(sbox_file)
    ws = wb[worksheet]

    s = []
    for row in range(1,17):
        for col in 'ABCDEFGHIJKLMNOP':
            s.append(int(ws[col+str(row)].value))

    return tuple(s)

def kdf(key, size=32):
    salt = b'tYa7_l^6Uz!@*jah*->K)+%@eEG6$.l'

    derived_bytes = pyscrypt.hash(key, salt, 1024, 1, 1, size)

    return derived_bytes[:16], derived_bytes[16:]

def pad(s):
    rem = 16 - (len(s) % 16)
    return s + bytes([rem]) * rem

def ecb_encrypt(plaintext, key, sbox):
    ciphertext = b''
    aes = AES(key, sbox=sbox)

    blocks = [plaintext[i:i+16] for i in range(0, len(plaintext), 16)]
    for block in blocks:
        ciphertext += aes.encrypt_block(block)

    return ciphertext

def ecb_original_encrypt(plaintext, key):
    ciphertext = b''
    aes = aesoriginal.AES(key)

    blocks = [plaintext[i:i+16] for i in range(0, len(plaintext), 16)]
    for block in blocks:
        ciphertext += aes.encrypt_block(block)

    return ciphertext
""""
def cbc_encrypt(plaintext, key, iv):
    ciphertext = b''
    aes = AES(key, sbox=SBOX)

    ciphertext = aes.encrypt_cbc(plaintext, iv)

    return ciphertext

def cbc_original_encrypt(plaintext, key, iv):
    ciphertext = b''
    aes = aesoriginal.AES(key)

    ciphertext = aes.encrypt_cbc(plaintext, iv)

    return ciphertext

def ctr_encrypt(plaintext, key, iv):
    ciphertext = b''
    aes = AES(key, sbox=SBOX)

    ciphertext = aes.encrypt_ctr(plaintext, iv)

    return ciphertext

def ctr_original_encrypt(plaintext, key, iv):
    ciphertext = b''
    aes = aesoriginal.AES(key)

    ciphertext = aes.encrypt_ctr(plaintext, iv)

    return ciphertext

def ofb_encrypt(plaintext, key, iv):
    ciphertext = b''
    aes = AES(key, sbox=SBOX)

    ciphertext = aes.encrypt_ofb(plaintext, iv)

    return ciphertext

def ofb_original_encrypt(plaintext, key, iv):
    ciphertext = b''
    aes = aesoriginal.AES(key)

    ciphertext = aes.encrypt_ofb(plaintext, iv)

    return ciphertext

def cfb_encrypt(plaintext, key, iv):
    ciphertext = b''
    aes = AES(key, sbox=SBOX)

    ciphertext = aes.encrypt_cfb(plaintext, iv)

    return ciphertext

def cfb_original_encrypt(plaintext, key, iv):
    ciphertext = b''
    aes = aesoriginal.AES(key)

    ciphertext = aes.encrypt_cfb(plaintext, iv)

    return ciphertext

def ecb_decrypt(ciphertext, key):
    plaintext = b''
    aes = AES(key, sbox=SBOX)

    blocks = [ciphertext[i:i+16] for i in range(0, len(ciphertext), 16)]
    for block in blocks:
        plaintext += aes.decrypt_block(block)

    return plaintext

def cbc_decrypt(ciphertext, key, iv):
    plaintext = b''
    aes = AES(key, sbox=SBOX)

    plaintext = aes.decrypt_cbc(ciphertext, iv)

    return plaintext

def ctr_decrypt(ciphertext, key, iv):
    plaintext = b''
    aes = AES(key, sbox=SBOX)

    plaintext = aes.decrypt_ctr(ciphertext, iv)

    return plaintext

def ofb_decrypt(ciphertext, key, iv):
    plaintext = b''
    aes = AES(key, sbox=SBOX)

    plaintext = aes.decrypt_ofb(ciphertext, iv)

    return plaintext

def cfb_decrypt(ciphertext, key, iv):
    plaintext = b''
    aes = AES(key, sbox=SBOX)

    plaintext = aes.decrypt_cfb(ciphertext, iv)

    return plaintext
"""

#STATIC FILES
@app.route('/<path:path>')
def send_static(path):
    return send_from_directory('static/', path)

#PUBLIC FILES
@app.route('/download/<path:filename>', methods=['GET'])
def download(filename):
    path = 'tmp/'
    return send_file(path+filename, as_attachment=True)

#API ENDPOINT
@app.route('/api/encrypt', methods=["POST"])
def encrypt_func():
    response = {}

    key = request.form.get('key')
    size = request.form.get('size')
    mode = request.form.get('mode')

    if size == '128':
        size = 16
    elif size == '192':
        size = 24
    else:
        size = 32

    iv, cipher_key = kdf(str.encode(key), size+16)
    
    if request.files:
        image = request.files["image"]
        try:
            original_img = Image.open(image)
        except Exception as e:
            return json.dumps({"status":"error","msg":str(e)})
        rgb, image_size, image_mode = img_to_rgb(image)
        plaintext = rgb_to_bytes(rgb)

        cipher_key_i = cipher_key[:-1]
        cipher_key_i += bytes([cipher_key[-1] ^ 1])

        plaintext_i = plaintext[1:]
        plaintext_i = bytes([plaintext[0] ^ 1]) + plaintext_i
        
        try:
            sbox_file = request.files["sbox-file"]
            sbox_sheet = request.form.get('sbox')

            SBOX = read_excel(sbox_file, sbox_sheet)
            # print(SBOX)
            # print(image, key, size, mode, sbox_file, sbox_sheet)
            if len(SBOX) != 256:
                return json.dumps({"status":"error", "msg": "SBOX tidak valid"})
        except Exception as e:
            return json.dumps({"status":"error","msg":"SBOX tidak valid: "+str(e)})
        

        if mode == "cbc":
            start_time = time.time()
            ciphertext1 = cbc_encrypt(plaintext, cipher_key, iv)
            run_time1 = time.time() - start_time

            start_time = time.time()
            ciphertext2 = cbc_original_encrypt(plaintext, cipher_key, iv)
            run_time2 = time.time() - start_time

            ciphertext1_i_key = cbc_encrypt(plaintext, cipher_key_i, iv)
            ciphertext1_i_plaintext = cbc_encrypt(plaintext_i, cipher_key, iv)
            ciphertext2_i_key = cbc_original_encrypt(plaintext, cipher_key_i, iv)
            ciphertext2_i_plaintext = cbc_original_encrypt(plaintext_i, cipher_key, iv)

        elif mode == "ctr":
            start_time = time.time()
            ciphertext1 = ctr_encrypt(plaintext, cipher_key, iv)
            run_time1 = time.time() - start_time

            start_time = time.time()
            ciphertext2 = ctr_original_encrypt(plaintext, cipher_key, iv)
            run_time2 = time.time() - start_time

            ciphertext1_i_key = ctr_encrypt(plaintext, cipher_key_i, iv)
            ciphertext1_i_plaintext = ctr_encrypt(plaintext_i, cipher_key, iv)
            ciphertext2_i_key = ctr_original_encrypt(plaintext, cipher_key_i, iv)
            ciphertext2_i_plaintext = ctr_original_encrypt(plaintext_i, cipher_key, iv)

        elif mode == "ofb":
            start_time = time.time()
            ciphertext1 = ofb_encrypt(plaintext, cipher_key, iv)
            run_time1 = time.time() - start_time

            start_time = time.time()
            ciphertext2 = ofb_original_encrypt(plaintext, cipher_key, iv)
            run_time2 = time.time() - start_time

            ciphertext1_i_key = ofb_encrypt(plaintext, cipher_key_i, iv)
            ciphertext1_i_plaintext = ofb_encrypt(plaintext_i, cipher_key, iv)
            ciphertext2_i_key = ofb_original_encrypt(plaintext, cipher_key_i, iv)
            ciphertext2_i_plaintext = ofb_original_encrypt(plaintext_i, cipher_key, iv)

        elif mode == "cfb":
            start_time = time.time()
            ciphertext1 = cfb_encrypt(plaintext, cipher_key, iv)
            run_time1 = time.time() - start_time

            start_time = time.time()
            ciphertext2 = cfb_original_encrypt(plaintext, cipher_key, iv)
            run_time2 = time.time() - start_time

            ciphertext1_i_key = cfb_encrypt(plaintext, cipher_key_i, iv)
            ciphertext1_i_plaintext = cfb_encrypt(plaintext_i, cipher_key, iv)
            ciphertext2_i_key = cfb_original_encrypt(plaintext, cipher_key_i, iv)
            ciphertext2_i_plaintext = cfb_original_encrypt(plaintext_i, cipher_key, iv)

        else:
            start_time = time.time()
            ciphertext1 = ecb_encrypt(plaintext, cipher_key, SBOX)
            run_time1 = time.time() - start_time

            start_time = time.time()
            ciphertext2 = ecb_original_encrypt(plaintext, cipher_key)
            run_time2 = time.time() - start_time

            ciphertext1_i_key = ecb_encrypt(plaintext, cipher_key_i, SBOX)
            ciphertext1_i_plaintext = ecb_encrypt(plaintext_i, cipher_key, SBOX)
            ciphertext2_i_key = ecb_original_encrypt(plaintext, cipher_key_i)
            ciphertext2_i_plaintext = ecb_original_encrypt(plaintext_i, cipher_key)

        img1 = bytes_to_image(ciphertext1, image_mode, image_size)
        img2 = bytes_to_image(ciphertext2, image_mode, image_size)

        img1_i_key = bytes_to_image(ciphertext1_i_key, image_mode, image_size)
        img1_i_plaintext = bytes_to_image(ciphertext1_i_plaintext, image_mode, image_size)
        img2_i_key = bytes_to_image(ciphertext2_i_key, image_mode, image_size)
        img2_i_plaintext = bytes_to_image(ciphertext2_i_plaintext, image_mode, image_size)

        img1.save("tmp/ciphertext_modified_aes.png", "PNG")
        img2.save("tmp/ciphertext_original_aes.png", "PNG")

        encrypted_image = image_to_base64(img1)
        encrypted_image_original = image_to_base64(img2)
        image_byte = image_to_base64(original_img)

        

        entropy = [
            image_entropy(img1, image_mode),
            image_entropy(img2, image_mode),
            image_entropy(original_img, image_mode)
        ]

        if image_mode == 'RGB':
            hist1 = img1.histogram()
            hist2 = img2.histogram()
            hist3 = original_img.histogram()
            histogram = [
                [hist1[:256], hist1[256:512], hist1[512:]],
                [hist2[:256], hist2[256:512], hist2[512:]],
                [hist3[:256], hist3[256:512], hist3[512:]]
            ]

            diff1 = [0,0,0]
            diff2 = [0,0,0]
            for i in range(256):
                if hist1[:256][i] != hist3[:256][i]:
                    diff1[0] += 1
                if hist1[256:512][i] != hist3[256:512][i]:
                    diff1[1] += 1
                if hist1[512:][i] != hist3[512:][i]:
                    diff1[2] += 1

                if hist2[:256][i] != hist3[:256][i]:
                    diff2[0] += 1
                if hist2[256:512][i] != hist3[256:512][i]:
                    diff2[1] += 1
                if hist2[512:][i] != hist3[512:][i]:
                    diff2[2] += 1
        else:
            hist1 = img1.histogram()
            hist2 = img2.histogram()
            hist3 = original_img.histogram()
            histogram = [
                [hist1],
                [hist2],
                [hist3]
            ]

            diff1 = [0]
            diff2 = [0]
            for i in range(256):
                if hist1[i] != hist3[i]:
                    diff1[0] += 1
                if hist2[i] != hist3[i]:
                    diff2[0] += 1

        uaci = {
            "modified": {
                "key": UACI(img1, img1_i_key),
                "plaintext": UACI(img1, img1_i_plaintext)
            },
            "original": {
                "key": UACI(img2, img2_i_key),
                "plaintext": UACI(img2, img2_i_plaintext)
            }
        }

        npcr = {
            "modified": {
                "key": NPCR(img1, img1_i_key),
                "plaintext": NPCR(img1, img1_i_plaintext)
            },
            "original": {
                "key": NPCR(img2, img2_i_key),
                "plaintext": NPCR(img2, img2_i_plaintext)
            }
        }

        response = {
            "status":"ok",
            "plaintext": image_byte,
            "ciphertext":[encrypted_image, encrypted_image_original],
            "entropy": entropy,
            "histogram": histogram,
            "uaci": uaci,
            "npcr": npcr,
            "runtime": [run_time1, run_time2],
            "diff": [diff1, diff2]
        }
    else:
        response = {"status":"error","msg":"Gambar tidak ditemukan"}

    return json.dumps(response)

@app.route('/api/decrypt', methods=["POST"])
def decrypt_func():
    response = {}

    key = request.form.get('key')
    size = request.form.get('size')
    mode = request.form.get('mode')

    if size == '128':
        size = 16
    elif size == '192':
        size = 24
    else:
        size = 32

    iv, cipher_key = kdf(str.encode(key), size+16)
    
    if request.files:
        image = request.files["image"]
        original_img = Image.open(image)
        rgb, image_size, image_mode = img_to_rgb(image)
        ciphertext = rgb_to_bytes(rgb)

        try:
            df_m1 = request.files["sbox"]
            SBOX = read_excel(df_m1)
            if len(SBOX) != 256:
                return json.dumps({"status":"error", "msg": "SBOX tidak valid"})
        except Exception as e:
            return json.dumps({"status":"error","msg":str(e)})

        if mode == "cbc":
            plaintext1 = cbc_decrypt(ciphertext, cipher_key, iv)
        elif mode == "ctr":
            plaintext1 = ctr_decrypt(ciphertext, cipher_key, iv)
        elif mode == "ofb":
            plaintext1 = ofb_decrypt(ciphertext, cipher_key, iv)
        elif mode == "cfb":
            plaintext1 = cfb_decrypt(ciphertext, cipher_key, iv)
        else:
            plaintext1 = ecb_decrypt(ciphertext, cipher_key)

        
        img1 = bytes_to_image(plaintext1, image_mode, image_size)
        img1.save("tmp/plaintext.png", "PNG")
        decrypted_image = image_to_base64(img1)

        response = {
            "status":"ok",
            "plaintext":[decrypted_image]
        }
    else:
        response = {"status":"error","msg":"Gambar tidak ditemukan"}

    return json.dumps(response)

#VIEWS

@app.route('/')
def homepage():
    return render_template("index.html")

@app.route('/encrypt')
def encrypt_page():
    return render_template("encrypt.html")

@app.route('/decrypt')
def decrypt_page():
    return render_template("decrypt.html")

if __name__ == "__main__":
    app.run(debug=True, port=8000)